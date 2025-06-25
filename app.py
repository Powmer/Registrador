import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

excel_file = ""
carrinho = []
comandas = {}

de1a100 = list(range(1, 101))
de50em100 = [50 * i for i in range(101)]

# ---------------- Funções Excel e edição ----------------

def criar_excel():
    global excel_file
    if not excel_file:
        return
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendas"
        ws.append(["Comanda/Carrinho", "Data e Hora", "Produto", "Quantidade", "Preço", "Entrega", "Pagamento"])
        wb.save(excel_file)

def selecionar_diretorio():
    global excel_file
    pasta = filedialog.askdirectory(title="Selecione o diretório para salvar o arquivo Excel")
    if pasta:
        excel_file = os.path.join(pasta, "vendas.xlsx")
        criar_excel()
        messagebox.showinfo("Sucesso", f"Arquivo criado em: {excel_file}")
def selecionar_arquivo_excel():
    global excel_file
    arquivo = filedialog.askopenfilename(title="Selecione um arquivo Excel", filetypes=[("Arquivos Excel", "*.xlsx")])
    if arquivo:
        excel_file = arquivo
        messagebox.showinfo("Arquivo Carregado", f"Arquivo carregado: {excel_file}")
def carregar_planilha():
    global excel_file
    if not excel_file or not os.path.exists(excel_file):
        messagebox.showwarning("Erro", "Nenhum arquivo Excel carregado ou arquivo não existe.")
        return

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao abrir arquivo: {e}")
        return

    for i in tree_planilha.get_children():
        tree_planilha.delete(i)

    # Mostrar dados no treeview
    for row in ws.iter_rows(min_row=2, values_only=False):
        valores = []
        for cell in row:
            valores.append(cell.value)
        tree_planilha.insert("", "end", values=valores)

def salvar_planilha():
    global excel_file
    if not excel_file:
        messagebox.showwarning("Erro", "Nenhum arquivo Excel carregado.")
        return

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.delete_rows(2, ws.max_row)  #Reset

    for item_id in tree_planilha.get_children():
        valores = tree_planilha.item(item_id, "values")
        ws.append(valores)

    try:
        wb.save(excel_file)
        messagebox.showinfo("Sucesso", "Planilha salva com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar planilha: {e}")

def excluir_linha_planilha():
    selecionados = tree_planilha.selection()
    if not selecionados:
        messagebox.showwarning("Aviso", "Selecione uma ou mais linhas para excluir.")
        return
    for sel in selecionados:
        tree_planilha.delete(sel)


def on_tree_double_click(event):
    item = tree_planilha.identify_row(event.y)
    column = tree_planilha.identify_column(event.x)
    if not item or not column:
        return

    col_index = int(column.replace("#","")) - 1
    if col_index == 1:  
        return

    x, y, width, height = tree_planilha.bbox(item, column)
    valor_atual = tree_planilha.set(item, column)

    entry_edit = tk.Entry(tree_planilha)
    entry_edit.place(x=x, y=y, width=width, height=height)
    entry_edit.insert(0, valor_atual)
    entry_edit.focus_set()

    def salvar_edicao(event=None):
        novo_valor = entry_edit.get()
      
        if col_index == 3:
            try:
                if "." in novo_valor:
                    novo_valor = float(novo_valor)
                else:
                    novo_valor = int(novo_valor)
                if novo_valor < 0:
                    raise ValueError
            except:
                messagebox.showwarning("Valor inválido", "Quantidade deve ser um número não negativo.")
                entry_edit.focus_set()
                return
        tree_planilha.set(item, column, novo_valor)
        entry_edit.destroy()

    entry_edit.bind("<Return>", salvar_edicao)
    entry_edit.bind("<FocusOut>", lambda e: entry_edit.destroy())

# ---------------- Funções Carrinho ----------------

def calcular_preco(produto, quantidade):
    if produto == "Combo Individual":
        return round(quantidade * 29.99, 2)
    elif produto == "Combo Família":
        return round(quantidade * 79.99, 2)
    elif produto == "Kilo":
        return round((quantidade / 1000) * 89.99, 2)
    return 0.0

def adicionar_ao_carrinho():
    produto = product_type_var.get()
    entrega = delivery_type_var.get()
    if produto == "Kilo":
        quantidade = quantidade_gramas_var.get()
    else:
        quantidade = quantidade_var.get()

    if not produto or not entrega or quantidade <= 0:
        messagebox.showwarning("Atenção", "Preencha produto, quantidade e entrega.")
        return

    preco = calcular_preco(produto, quantidade)
    carrinho.append({
        "produto": produto,
        "quantidade": quantidade,
        "preco": preco,
        "entrega": entrega
    })

    atualizar_lista_carrinho()
    atualizar_total_carrinho()
    limpar_campos_venda()

def atualizar_lista_carrinho():
    for i in lista_carrinho.get_children():
        lista_carrinho.delete(i)
    for item in carrinho:
        lista_carrinho.insert("", "end", values=(item["produto"], item["quantidade"], f"R$ {item['preco']}", item["entrega"]))

def atualizar_total_carrinho():
    total = sum(item["preco"] for item in carrinho)
    preco_total_carrinho_var.set(f"R$ {round(total, 2)}")

def registrar_carrinho():
    global excel_file
    if not excel_file:
        messagebox.showwarning("Atenção", "Selecione ou crie um arquivo Excel antes.")
        return
    pagamento = payment_method_var.get()
    if not pagamento:
        messagebox.showwarning("Atenção", "Informe o método de pagamento.")
        return
    if not carrinho:
        messagebox.showwarning("Atenção", "Carrinho vazio.")
        return

    data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    for item in carrinho:
        ws.append(["Carrinho", data_hora, item["produto"], item["quantidade"], item["preco"], item["entrega"], pagamento])
    wb.save(excel_file)
    messagebox.showinfo("Sucesso", "Venda do carrinho registrada!")
    carrinho.clear()
    atualizar_lista_carrinho()
    atualizar_total_carrinho()

def excluir_item_carrinho():
    selecionados = lista_carrinho.selection()
    if not selecionados:
        messagebox.showwarning("Aviso", "Selecione um ou mais itens para excluir.")
        return

    for selecionado in selecionados:
        index = lista_carrinho.index(selecionado)
        lista_carrinho.delete(selecionado)
        del carrinho[index]

    atualizar_total_carrinho()

def limpar_campos_venda():
    product_type_var.set("")
    quantidade_var.set(1)
    quantidade_gramas_var.set(0)
    delivery_type_var.set("")

def atualizar_campos_quantidade(*args):
    produto = product_type_var.get()
    if produto == "Kilo":
        quantidade_label.grid_remove()
        quantidade_spin.grid_remove()
        quantidade_gramas_label.grid(row=2, column=0)
        quantidade_gramas_spin.grid(row=2, column=1)
    else:
        quantidade_gramas_label.grid_remove()
        quantidade_gramas_spin.grid_remove()
        quantidade_label.grid(row=1, column=0)
        quantidade_spin.grid(row=1, column=1)

# ---------- Interface gráfica ----------

root = tk.Tk()
root.title("Sistema de Vendas e Comandas")
root.option_add("*Font", "Helvetica 10")
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview.Heading", font=("Helvetica", 10, "bold"))
style.configure("TButton", padding=5)

# Frame de Vendas
frame_venda = ttk.LabelFrame(root, text="Venda")
frame_venda.pack(side="left", padx=10, pady=10, fill="both")

ttk.Label(frame_venda, text="Produto").grid(row=0, column=0)
product_type_var = tk.StringVar()
product_combobox = ttk.Combobox(frame_venda, textvariable=product_type_var, values=["Combo Individual", "Combo Família", "Kilo"])
product_combobox.grid(row=0, column=1)

quantidade_label = ttk.Label(frame_venda, text="Qtd Produtos")
quantidade_label.grid(row=1, column=0)
quantidade_var = tk.IntVar(value=1)
quantidade_spin = ttk.Spinbox(frame_venda, from_=1, to=100, textvariable=quantidade_var, width=5)
quantidade_spin.grid(row=1, column=1)

quantidade_gramas_label = ttk.Label(frame_venda, text="Qtd (g)")
quantidade_gramas_var = tk.IntVar(value=0)
quantidade_gramas_spin = ttk.Spinbox(frame_venda, from_=0, to=5000, increment=50, textvariable=quantidade_gramas_var, width=7)
quantidade_gramas_label.grid_remove()
quantidade_gramas_spin.grid_remove()

ttk.Label(frame_venda, text="Entrega").grid(row=3, column=0)
delivery_type_var = tk.StringVar()
ttk.Combobox(frame_venda, textvariable=delivery_type_var, values=["Retirada", "Entrega"]).grid(row=3, column=1)

ttk.Button(frame_venda, text="Adicionar ao Carrinho", command=adicionar_ao_carrinho).grid(row=4, column=0, columnspan=2, pady=5)

ttk.Label(frame_venda, text="Pagamento").grid(row=5, column=0)
payment_method_var = tk.StringVar()
ttk.Combobox(frame_venda, textvariable=payment_method_var, values=["Pix", "Cartão", "Dinheiro"]).grid(row=5, column=1)

ttk.Button(frame_venda, text="Finalizar Venda", command=registrar_carrinho).grid(row=6, column=0, columnspan=2, pady=5)
ttk.Button(frame_venda, text="Selecionar Pasta", command=selecionar_diretorio).grid(row=7, column=0, columnspan=2, pady=5)
ttk.Button(frame_venda, text="Carregar Arquivo Excel", command=lambda: [selecionar_arquivo_excel(), carregar_planilha()]).grid(row=8, column=0, columnspan=2, pady=5)

product_type_var.trace_add("write", atualizar_campos_quantidade)
atualizar_campos_quantidade()

# Frame Carrinho
frame_carrinho = ttk.LabelFrame(root, text="Carrinho")
frame_carrinho.pack(side="left", pady=10, fill="both")

lista_carrinho = ttk.Treeview(frame_carrinho, columns=("Produto", "Qtd", "Preço", "Entrega"), show="headings", height=7)

lista_carrinho.heading("Produto", text="Produto")
lista_carrinho.heading("Qtd", text="Qtd")
lista_carrinho.heading("Preço", text="Preço")
lista_carrinho.heading("Entrega", text="Entrega")

lista_carrinho.column("Produto", width=150)
lista_carrinho.column("Qtd", width=50)
lista_carrinho.column("Preço", width=70)
lista_carrinho.column("Entrega", width=100)
for col in ("Produto", "Qtd", "Preço", "Entrega"):
    lista_carrinho.heading(col, text=col)
lista_carrinho.pack(padx=5, pady=5)

ttk.Label(frame_carrinho, text="Total:").pack()
preco_total_carrinho_var = tk.StringVar(value="R$ 0.0")
ttk.Label(frame_carrinho, textvariable=preco_total_carrinho_var, font=("Helvetica", 10, "bold")).pack()
ttk.Button(frame_carrinho, text="Excluir Item(s)", command=excluir_item_carrinho).pack(pady=5)

# Frame Planilha Excel 
frame_planilha = ttk.LabelFrame(root, text="Planilha Excel")
frame_planilha.pack(side="right", padx=10, pady=10, fill="both", expand=True)

colunas_planilha = ["Comanda/Carrinho", "Data e Hora", "Produto", "Quantidade", "Preço", "Entrega", "Pagamento"]

tree_planilha = ttk.Treeview(frame_planilha, columns=colunas_planilha, show="headings")
for c in colunas_planilha:
    tree_planilha.heading(c, text=c)
    tree_planilha.column(c, width=100)
tree_planilha.pack(fill="both", expand=True)

tree_planilha.bind("<Double-1>", on_tree_double_click)

btn_frame = ttk.Frame(frame_planilha)
btn_frame.pack(fill="x")

ttk.Button(btn_frame, text="Excluir Linha(s)", command=excluir_linha_planilha).pack(side="left", padx=5, pady=5)
ttk.Button(btn_frame, text="Salvar Alterações", command=salvar_planilha).pack(side="left", padx=5, pady=5)

root.mainloop()
