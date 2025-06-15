from kivy.app import App
from kivy.lang import Builder
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserIconView
from kivy.properties import StringProperty, ListProperty, NumericProperty
from kivy.uix.spinner import Spinner
from kivy.uix.textinput import TextInput
from kivy.uix.recycleview import RecycleView
from kivy.uix.button import Button
from kivy.clock import mainthread
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

KV = '''
<EditableLabel@TextInput>:
    multiline: False
    input_filter: 'float' if self.is_number else None
    readonly: False
    is_number: False

<RootWidget>:
    orientation: "vertical"
    padding: 10
    spacing: 10

    BoxLayout:
        size_hint_y: None
        height: "40dp"
        spacing: 10

        Spinner:
            id: product_spinner
            text: "Produto"
            values: ["Combo Individual", "Combo Família", "Kilo"]
            on_text: root.on_product_select(self.text)

        BoxLayout:
            id: quantidade_box
            size_hint_x: 0.3
            TextInput:
                id: quantidade_input
                hint_text: "Qtd Produtos"
                input_filter: "int"
                multiline: False
                on_text_validate: root.on_quantidade_change(self.text)
            # Será trocado para gramas se necessário

        Spinner:
            id: entrega_spinner
            text: "Entrega"
            values: ["Retirada", "Entrega"]

    BoxLayout:
        size_hint_y: None
        height: "40dp"
        spacing: 10

        Spinner:
            id: pagamento_spinner
            text: "Pagamento"
            values: ["Pix", "Cartão", "Dinheiro"]

        Button:
            text: "Adicionar ao Carrinho"
            on_press: root.adicionar_ao_carrinho()

        Button:
            text: "Finalizar Venda"
            on_press: root.registrar_carrinho()

    BoxLayout:
        size_hint_y: None
        height: "40dp"
        spacing: 10

        Button:
            text: "Selecionar Pasta"
            on_press: root.selecionar_diretorio()

        Button:
            text: "Carregar Arquivo Excel"
            on_press: root.selecionar_arquivo_excel()

    Label:
        id: total_label
        text: "Total: R$ 0.0"
        size_hint_y: None
        height: "30dp"

    Label:
        text: "Carrinho (toque para excluir item):"
        size_hint_y: None
        height: "30dp"

    RecycleView:
        id: rv
        viewclass: 'Button'
        data: root.rv_data
        size_hint_y: 0.4
        scroll_type: ['bars', 'content']
        bar_width: 10
        RecycleBoxLayout:
            default_size: None, dp(40)
            default_size_hint: 1, None
            size_hint_y: None
            height: self.minimum_height
            orientation: 'vertical'
'''

class RootWidget(BoxLayout):
    excel_file = StringProperty("")
    carrinho = ListProperty([])
    rv_data = ListProperty([])
    total = NumericProperty(0.0)

    def on_product_select(self, produto):
        quantidade_box = self.ids.quantidade_box
        quantidade_box.clear_widgets()
        if produto == "Kilo":
            # entrada para gramas
            ti = TextInput(hint_text="Qtd (g)", input_filter="int", multiline=False)
        else:
            # entrada para quantidade inteira
            ti = TextInput(hint_text="Qtd Produtos", input_filter="int", multiline=False)
        ti.bind(text=self.on_quantidade_change)
        quantidade_box.add_widget(ti)
        self.quantidade_input = ti

    def on_quantidade_change(self, instance, valor):
        # Aqui pode validar ou formatar se quiser
        pass

    def adicionar_ao_carrinho(self):
        produto = self.ids.product_spinner.text
        entrega = self.ids.entrega_spinner.text
        pagamento = self.ids.pagamento_spinner.text
        if produto not in ["Combo Individual", "Combo Família", "Kilo"]:
            self.show_popup("Erro", "Selecione um produto válido.")
            return
        if entrega not in ["Retirada", "Entrega"]:
            self.show_popup("Erro", "Selecione o tipo de entrega.")
            return
        if pagamento not in ["Pix", "Cartão", "Dinheiro"]:
            self.show_popup("Erro", "Selecione o método de pagamento.")
            return
        try:
            quantidade_str = self.quantidade_input.text
            quantidade = int(quantidade_str)
        except:
            self.show_popup("Erro", "Quantidade inválida.")
            return
        if quantidade <= 0:
            self.show_popup("Erro", "Quantidade deve ser maior que zero.")
            return

        preco = self.calcular_preco(produto, quantidade)

        item = {
            "produto": produto,
            "quantidade": quantidade,
            "preco": preco,
            "entrega": entrega,
            "pagamento": pagamento,
        }

        self.carrinho.append(item)
        self.atualizar_lista_carrinho()
        self.atualizar_total_carrinho()
        self.limpar_campos()

    def calcular_preco(self, produto, quantidade):
        if produto == "Combo Individual":
            return round(quantidade * 29.99, 2)
        elif produto == "Combo Família":
            return round(quantidade * 79.99, 2)
        elif produto == "Kilo":
            return round((quantidade / 1000) * 89.99, 2)
        return 0.0

    def atualizar_lista_carrinho(self):
        data = []
        for i, item in enumerate(self.carrinho):
            display_text = f"{i+1}. {item['produto']} - Qtd: {item['quantidade']} - R$ {item['preco']} - {item['entrega']}"
            data.append({
                'text': display_text,
                'on_press': lambda x=i: self.excluir_item(x)
            })
        self.rv_data = data

    def excluir_item(self, index):
        if 0 <= index < len(self.carrinho):
            del self.carrinho[index]
            self.atualizar_lista_carrinho()
            self.atualizar_total_carrinho()

    def atualizar_total_carrinho(self):
        total = sum(item['preco'] for item in self.carrinho)
        self.total = round(total, 2)
        self.ids.total_label.text = f"Total: R$ {self.total}"

    def limpar_campos(self):
        self.ids.product_spinner.text = "Produto"
        self.ids.entrega_spinner.text = "Entrega"
        self.ids.pagamento_spinner.text = "Pagamento"
        self.quantidade_input.text = ""

    def selecionar_diretorio(self):
        content = FileChooserIconView(path='.', filters=['*.xlsx'])
        popup = Popup(title="Selecione o diretório (só para salvar)", content=content, size_hint=(0.9, 0.9))
        content.bind(on_submit=lambda instance, selection, touch: self._salvar_diretorio(selection, popup))
        popup.open()

    def _salvar_diretorio(self, selection, popup):
        if selection:
            pasta = selection[0]
            if not os.path.isdir(pasta):
                pasta = os.path.dirname(pasta)
            self.excel_file = os.path.join(pasta, "vendas.xlsx")
            if not os.path.exists(self.excel_file):
                wb = Workbook()
                ws = wb.active
                ws.title = "Vendas"
                ws.append(["Comanda/Carrinho", "Data e Hora", "Produto", "Quantidade", "Preço", "Entrega", "Pagamento"])
                wb.save(self.excel_file)
            self.show_popup("Sucesso", f"Arquivo criado em:\n{self.excel_file}")
            popup.dismiss()

    def selecionar_arquivo_excel(self):
        content = FileChooserIconView(path='.', filters=['*.xlsx'])
        popup = Popup(title="Selecione o arquivo Excel", content=content, size_hint=(0.9, 0.9))
        content.bind(on_submit=lambda instance, selection, touch: self._carregar_arquivo_excel(selection, popup))
        popup.open()

    def _carregar_arquivo_excel(self, selection, popup):
        if selection:
            arquivo = selection[0]
            self.excel_file = arquivo
            self.carregar_planilha()
            self.show_popup("Arquivo Carregado", f"Arquivo carregado:\n{arquivo}")
            popup.dismiss()

    def carregar_planilha(self):
        if not self.excel_file or not os.path.exists(self.excel_file):
            return
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb.active
        self.planilha_dados = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            self.planilha_dados.append(list(row))
        self.mostrar_planilha_popup()

    def mostrar_planilha_popup(self):
        layout = BoxLayout(orientation='vertical')
        header = ["Comanda/Carrinho", "Data e Hora", "Produto", "Quantidade", "Preço", "Entrega", "Pagamento"]

        # Cria header fixo
        header_box = BoxLayout(size_hint_y=None, height=30)
        for h in header:
            header_box.add_widget(Button(text=h, background_color=(0.7,0.7,0.7,1), disabled=True))
        layout.add_widget(header_box)

        # Área para linhas editáveis
        from kivy.uix.scrollview import ScrollView
        scroll = ScrollView(size_hint=(1, 0.8))
        linhas_box = BoxLayout(orientation='vertical', size_hint_y=None)
        linhas_box.bind(minimum_height=linhas_box.setter('height'))

        self.inputs = []

        for i, linha in enumerate(self.planilha_dados):
            linha_box = BoxLayout(size_hint_y=None, height=30)
            linha_inputs = []
            for j, valor in enumerate(linha):
                ti = TextInput(text=str(valor), multiline=False)
                if j == 3 or j == 4:  # Quantidade e preço (numéricos)
                    ti.input_filter = 'float'
                linha_box.add_widget(ti)
                linha_inputs.append(ti)
            linhas_box.add_widget(linha_box)
            self.inputs.append(linha_inputs)

        scroll.add_widget(linhas_box)
        layout.add_widget(scroll)

        btn_save = Button(text="Salvar Alterações", size_hint_y=None, height=40)
        btn_save.bind(on_press=self.salvar_alteracoes)
        layout.add_widget(btn_save)

        self.popup_planilha = Popup(title="Editar Planilha", content=layout, size_hint=(0.95, 0.95))
        self.popup_planilha.open()

    def salvar_alteracoes(self, *args):
        if not self.excel_file or not os.path.exists(self.excel_file):
            self.show_popup("Erro", "Arquivo Excel não carregado.")
            return

        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb.active

        # Limpa dados antigos (exceto cabeçalho)
        ws.delete_rows(2, ws.max_row)

        # Escreve novos dados editados
        for linha_inputs in self.inputs:
            valores = []
            for ti in linha_inputs:
                val = ti.text
                if val == "":
                    val = None
                valores.append(val)
            ws.append(valores)

        wb.save(self.excel_file)
        self.popup_planilha.dismiss()
        self.show_popup("Sucesso", "Alterações salvas com sucesso.")

    def registrar_carrinho(self):
        if not self.excel_file:
            self.show_popup("Erro", "Selecione ou crie um arquivo Excel antes.")
            return
        if not self.carrinho:
            self.show_popup("Erro", "Carrinho vazio.")
            return

        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb.active
        data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for item in self.carrinho:
            ws.append(["Carrinho", data_hora, item["produto"], item["quantidade"], item["preco"], item["entrega"], item["pagamento"]])

        wb.save(self.excel_file)
        self.carrinho.clear()
        self.atualizar_lista_carrinho()
        self.atualizar_total_carrinho()
        self.show_popup("Sucesso", "Venda do carrinho registrada!")

    def show_popup(self, title, message):
        popup = Popup(title=title,
                      content=Button(text=message, on_press=lambda x: popup.dismiss()),
                      size_hint=(0.6, 0.4))
        popup.open()


class AppVendas(App):
    def build(self):
        return RootWidget()


if __name__ == '__main__':
    AppVendas().run()
